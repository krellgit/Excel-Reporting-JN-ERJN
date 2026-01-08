#!/usr/bin/env python3
"""
Campaign Performance Report Generator
Reads CSV files directly and generates a complete Excel report.
No manual data pasting required - just run this script.
"""

import os
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import CellIsRule
import warnings
warnings.filterwarnings('ignore')

# ============================================================================
# CONFIGURATION - Update these paths if needed
# ============================================================================

CAMPAIGN_FILE = "Krelll_-Campaign_Test.csv"
BUSINESS_FILE = "BusinessReport- Sep2024 to Jan 2026.csv"
OUTPUT_FILE = "Campaign_Performance_Report.xlsx"

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

COLORS = {
    'primary': '1E3A5F',
    'secondary': '2E86AB',
    'jn': '2F5496',
    'non_jn': '0EA5E9',
    'branded': '22C55E',
    'competitor': 'F97316',
    'non_branded': '8B5CF6',
    'positive': '22C55E',
    'negative': 'EF4444',
    'light_gray': 'F3F4F6',
    'white': 'FFFFFF',
}

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

def apply_header_style(cell, color='primary'):
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color=COLORS[color], end_color=COLORS[color], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def format_currency(value):
    if pd.isna(value) or value == 0:
        return "$0"
    return f"${value:,.0f}"

def format_percent(value):
    if pd.isna(value):
        return "0.0%"
    return f"{value:.1f}%"

def format_decimal(value):
    if pd.isna(value):
        return "0.00"
    return f"{value:.2f}"

# ============================================================================
# DATA LOADING AND PROCESSING
# ============================================================================

def parse_currency(value):
    """Parse currency string to float."""
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # Remove $, commas, quotes
    cleaned = str(value).replace('$', '').replace(',', '').replace('"', '').strip()
    try:
        return float(cleaned) if cleaned else 0.0
    except:
        return 0.0

def parse_percent(value):
    """Parse percentage string to float."""
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace('%', '').replace('"', '').strip()
    try:
        return float(cleaned) if cleaned else 0.0
    except:
        return 0.0

def classify_portfolio(portfolio_name):
    """Classify portfolio as JN or Non-JN."""
    if pd.isna(portfolio_name):
        return 'Non-JN'
    return 'JN' if 'jn' in str(portfolio_name).lower() else 'Non-JN'

def classify_segment(campaign_name):
    """Classify campaign segment."""
    if pd.isna(campaign_name):
        return 'Non-Branded'
    name_lower = str(campaign_name).lower()
    if 'branded' in name_lower:
        return 'Branded'
    elif ' pat ' in name_lower or '- pat -' in name_lower or '_pat_' in name_lower:
        return 'Competitor'
    else:
        return 'Non-Branded'

def load_campaign_data(filepath):
    """Load and process campaign data."""
    print(f"Loading campaign data from {filepath}...")

    df = pd.read_csv(filepath, encoding='utf-8-sig')
    print(f"  Loaded {len(df):,} rows")
    print(f"  Columns: {list(df.columns)}")

    # Parse date - handle "Sep 01, 2024" format
    df['Date'] = pd.to_datetime(df['Date'], format='mixed', dayfirst=False)

    # Parse numeric columns
    df['Spend'] = df['Spend'].apply(parse_currency)
    df['Sales'] = df['7 Day Total Sales '].apply(parse_currency)
    df['Orders'] = df['7 Day Total Orders (#)'].apply(parse_currency)
    df['Impressions'] = df['Impressions'].apply(parse_currency)
    df['Clicks'] = df['Clicks'].apply(parse_currency)

    # Classify
    df['Portfolio_Type'] = df['Portfolio name'].apply(classify_portfolio)
    df['Segment'] = df['Campaign Name'].apply(classify_segment)

    # Add time dimensions
    df['Month'] = df['Date'].dt.to_period('M')
    df['Month_Label'] = df['Date'].dt.strftime('%b %Y')
    df['Week'] = df['Date'].dt.strftime('%Y-W%U')
    df['Year'] = df['Date'].dt.year

    print(f"  Date range: {df['Date'].min()} to {df['Date'].max()}")
    print(f"  Portfolio types: {df['Portfolio_Type'].value_counts().to_dict()}")
    print(f"  Segments: {df['Segment'].value_counts().to_dict()}")

    return df

def load_business_data(filepath):
    """Load and process business report data."""
    print(f"Loading business data from {filepath}...")

    df = pd.read_csv(filepath, encoding='utf-8-sig')
    print(f"  Loaded {len(df):,} rows")

    # Parse date - handle "9/1/24" format
    df['Date'] = pd.to_datetime(df['Date'], format='mixed', dayfirst=False)

    # Parse numeric columns
    df['Total_Sales'] = df['Ordered Product Sales'].apply(parse_currency)
    df['Units'] = df['Units Ordered'].apply(parse_currency)
    df['Sessions'] = df['Sessions - Total'].apply(parse_currency)

    # Add time dimensions
    df['Month'] = df['Date'].dt.to_period('M')
    df['Month_Label'] = df['Date'].dt.strftime('%b %Y')
    df['Week'] = df['Date'].dt.strftime('%Y-W%U')

    print(f"  Date range: {df['Date'].min()} to {df['Date'].max()}")

    return df

# ============================================================================
# AGGREGATION FUNCTIONS
# ============================================================================

def calc_metrics(df):
    """Calculate aggregated metrics from a dataframe."""
    spend = df['Spend'].sum()
    sales = df['Sales'].sum()
    orders = df['Orders'].sum()
    clicks = df['Clicks'].sum()
    impressions = df['Impressions'].sum()

    return {
        'Spend': spend,
        'Sales': sales,
        'Orders': orders,
        'Clicks': clicks,
        'Impressions': impressions,
        'ROAS': sales / spend if spend > 0 else 0,
        'ACoS': (spend / sales * 100) if sales > 0 else 0,
        'CVR': (orders / clicks * 100) if clicks > 0 else 0,
        'CPC': spend / clicks if clicks > 0 else 0,
        'CTR': (clicks / impressions * 100) if impressions > 0 else 0,
    }

def aggregate_by_month(campaign_df, business_df=None):
    """Aggregate data by month."""
    monthly = campaign_df.groupby('Month_Label').apply(
        lambda x: pd.Series(calc_metrics(x))
    ).reset_index()

    # Sort by actual date
    month_order = campaign_df.groupby('Month_Label')['Date'].min().sort_values()
    monthly['sort_order'] = monthly['Month_Label'].map({m: i for i, m in enumerate(month_order.index)})
    monthly = monthly.sort_values('sort_order').drop('sort_order', axis=1)

    # Add business data if available
    if business_df is not None:
        biz_monthly = business_df.groupby('Month_Label').agg({
            'Total_Sales': 'sum',
            'Units': 'sum',
            'Sessions': 'sum'
        }).reset_index()
        monthly = monthly.merge(biz_monthly, on='Month_Label', how='left')
        monthly['TACOS'] = (monthly['Spend'] / monthly['Total_Sales'] * 100).fillna(0)
        monthly['Organic_Sales'] = (monthly['Total_Sales'] - monthly['Sales']).clip(lower=0)

    return monthly

def aggregate_by_week(campaign_df):
    """Aggregate data by week."""
    weekly = campaign_df.groupby('Week').apply(
        lambda x: pd.Series(calc_metrics(x))
    ).reset_index()
    weekly = weekly.sort_values('Week')
    return weekly

def aggregate_by_segment(campaign_df):
    """Aggregate data by segment."""
    segments = campaign_df.groupby('Segment').apply(
        lambda x: pd.Series(calc_metrics(x))
    ).reset_index()
    return segments

def aggregate_by_portfolio(campaign_df):
    """Aggregate data by portfolio type."""
    portfolios = campaign_df.groupby('Portfolio_Type').apply(
        lambda x: pd.Series(calc_metrics(x))
    ).reset_index()
    return portfolios

def aggregate_by_portfolio_and_month(campaign_df):
    """Aggregate data by portfolio and month."""
    result = campaign_df.groupby(['Month_Label', 'Portfolio_Type']).apply(
        lambda x: pd.Series(calc_metrics(x))
    ).reset_index()
    return result

def aggregate_by_segment_and_month(campaign_df):
    """Aggregate data by segment and month."""
    result = campaign_df.groupby(['Month_Label', 'Segment']).apply(
        lambda x: pd.Series(calc_metrics(x))
    ).reset_index()
    return result

# ============================================================================
# EXCEL REPORT GENERATION
# ============================================================================

def create_summary_sheet(wb, campaign_df, business_df, monthly_data):
    """Create the Executive Summary sheet."""
    ws = wb.create_sheet("Executive Summary", 0)

    # Title
    ws['A1'] = "CAMPAIGN PERFORMANCE REPORT"
    ws['A1'].font = Font(bold=True, size=20, color=COLORS['primary'])
    ws.merge_cells('A1:H1')

    date_min = campaign_df['Date'].min().strftime('%b %d, %Y')
    date_max = campaign_df['Date'].max().strftime('%b %d, %Y')
    ws['A2'] = f"Date Range: {date_min} - {date_max}"
    ws['A2'].font = Font(size=12, color='6B7280')

    ws['A3'] = f"Generated: {datetime.now().strftime('%b %d, %Y %H:%M')}"
    ws['A3'].font = Font(size=10, italic=True, color='6B7280')

    # Overall Metrics
    overall = calc_metrics(campaign_df)
    total_sales = business_df['Total_Sales'].sum() if business_df is not None else 0
    tacos = (overall['Spend'] / total_sales * 100) if total_sales > 0 else 0

    ws['A5'] = "OVERALL PERFORMANCE"
    ws['A5'].font = Font(bold=True, size=14, color=COLORS['primary'])

    metrics = [
        ("Ad Spend", format_currency(overall['Spend'])),
        ("Ad Sales", format_currency(overall['Sales'])),
        ("ROAS", f"{overall['ROAS']:.2f}x"),
        ("ACoS", format_percent(overall['ACoS'])),
        ("Total Sales", format_currency(total_sales)),
        ("TACOS", format_percent(tacos)),
        ("Orders", f"{overall['Orders']:,.0f}"),
        ("Clicks", f"{overall['Clicks']:,.0f}"),
        ("CVR", format_percent(overall['CVR'])),
    ]

    row = 6
    for i, (label, value) in enumerate(metrics):
        col = (i % 3) * 2 + 1
        if i > 0 and i % 3 == 0:
            row += 2
        ws.cell(row=row, column=col, value=label).font = Font(size=9, color='6B7280')
        val_cell = ws.cell(row=row+1, column=col, value=value)
        val_cell.font = Font(bold=True, size=16)
        val_cell.alignment = Alignment(horizontal='left')

    # Portfolio Breakdown
    row = 14
    ws.cell(row=row, column=1, value="PORTFOLIO BREAKDOWN").font = Font(bold=True, size=14, color=COLORS['primary'])

    row += 1
    headers = ["Portfolio", "Spend", "Sales", "ROAS", "ACoS", "Orders"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        apply_header_style(cell)

    portfolio_data = aggregate_by_portfolio(campaign_df)
    for _, prow in portfolio_data.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=prow['Portfolio_Type']).font = Font(bold=True)
        ws.cell(row=row, column=2, value=prow['Spend']).number_format = '$#,##0'
        ws.cell(row=row, column=3, value=prow['Sales']).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=prow['ROAS']).number_format = '0.00'
        ws.cell(row=row, column=5, value=prow['ACoS']/100).number_format = '0.0%'
        ws.cell(row=row, column=6, value=prow['Orders']).number_format = '#,##0'

    # Segment Breakdown
    row += 3
    ws.cell(row=row, column=1, value="SEGMENT BREAKDOWN").font = Font(bold=True, size=14, color=COLORS['primary'])

    row += 1
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h.replace("Portfolio", "Segment"))
        apply_header_style(cell, 'branded')

    segment_data = aggregate_by_segment(campaign_df)
    for _, srow in segment_data.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=srow['Segment']).font = Font(bold=True)
        ws.cell(row=row, column=2, value=srow['Spend']).number_format = '$#,##0'
        ws.cell(row=row, column=3, value=srow['Sales']).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=srow['ROAS']).number_format = '0.00'
        ws.cell(row=row, column=5, value=srow['ACoS']/100).number_format = '0.0%'
        ws.cell(row=row, column=6, value=srow['Orders']).number_format = '#,##0'

    # Set column widths
    for i, w in enumerate([15, 14, 14, 10, 10, 12, 12, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def create_monthly_sheet(wb, monthly_data):
    """Create Monthly Performance sheet."""
    ws = wb.create_sheet("Monthly Performance")

    ws['A1'] = "MONTHLY PERFORMANCE"
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['primary'])

    # Headers
    headers = ["Month", "Spend", "Sales", "ROAS", "ACoS", "Orders", "Clicks", "CVR"]
    if 'Total_Sales' in monthly_data.columns:
        headers.extend(["Total Sales", "Organic", "TACOS"])

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        apply_header_style(cell)

    # Data
    for idx, row_data in monthly_data.iterrows():
        row = idx + 4
        ws.cell(row=row, column=1, value=row_data['Month_Label'])
        ws.cell(row=row, column=2, value=row_data['Spend']).number_format = '$#,##0'
        ws.cell(row=row, column=3, value=row_data['Sales']).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=row_data['ROAS']).number_format = '0.00'
        ws.cell(row=row, column=5, value=row_data['ACoS']/100).number_format = '0.0%'
        ws.cell(row=row, column=6, value=row_data['Orders']).number_format = '#,##0'
        ws.cell(row=row, column=7, value=row_data['Clicks']).number_format = '#,##0'
        ws.cell(row=row, column=8, value=row_data['CVR']/100).number_format = '0.00%'

        if 'Total_Sales' in monthly_data.columns:
            ws.cell(row=row, column=9, value=row_data['Total_Sales']).number_format = '$#,##0'
            ws.cell(row=row, column=10, value=row_data.get('Organic_Sales', 0)).number_format = '$#,##0'
            ws.cell(row=row, column=11, value=row_data.get('TACOS', 0)/100).number_format = '0.0%'

    # MoM Changes section
    start_row = len(monthly_data) + 6
    ws.cell(row=start_row, column=1, value="MONTH-OVER-MONTH CHANGES").font = Font(bold=True, size=14, color=COLORS['primary'])

    headers_mom = ["Month", "Spend %", "Sales %", "ROAS %"]
    for col, h in enumerate(headers_mom, 1):
        cell = ws.cell(row=start_row+1, column=col, value=h)
        apply_header_style(cell, 'secondary')

    for idx in range(1, len(monthly_data)):
        row = start_row + 2 + idx - 1
        curr = monthly_data.iloc[idx]
        prev = monthly_data.iloc[idx-1]

        ws.cell(row=row, column=1, value=curr['Month_Label'])

        spend_chg = ((curr['Spend'] - prev['Spend']) / prev['Spend'] * 100) if prev['Spend'] > 0 else 0
        sales_chg = ((curr['Sales'] - prev['Sales']) / prev['Sales'] * 100) if prev['Sales'] > 0 else 0
        roas_chg = ((curr['ROAS'] - prev['ROAS']) / prev['ROAS'] * 100) if prev['ROAS'] > 0 else 0

        for col, val in enumerate([spend_chg, sales_chg, roas_chg], 2):
            cell = ws.cell(row=row, column=col, value=val/100)
            cell.number_format = '0.0%'
            cell.font = Font(color=COLORS['positive'] if val >= 0 else COLORS['negative'])

    for i in range(1, 12):
        ws.column_dimensions[get_column_letter(i)].width = 14

def create_weekly_sheet(wb, weekly_data):
    """Create Weekly Performance sheet."""
    ws = wb.create_sheet("Weekly Performance")

    ws['A1'] = "WEEKLY PERFORMANCE"
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['primary'])

    headers = ["Week", "Spend", "Sales", "ROAS", "ACoS", "Orders", "Clicks"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        apply_header_style(cell)

    for idx, row_data in weekly_data.iterrows():
        row = idx + 4
        ws.cell(row=row, column=1, value=row_data['Week'])
        ws.cell(row=row, column=2, value=row_data['Spend']).number_format = '$#,##0'
        ws.cell(row=row, column=3, value=row_data['Sales']).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=row_data['ROAS']).number_format = '0.00'
        ws.cell(row=row, column=5, value=row_data['ACoS']/100).number_format = '0.0%'
        ws.cell(row=row, column=6, value=row_data['Orders']).number_format = '#,##0'
        ws.cell(row=row, column=7, value=row_data['Clicks']).number_format = '#,##0'

    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i)].width = 14

def create_segment_sheet(wb, campaign_df):
    """Create Segment Analysis sheet."""
    ws = wb.create_sheet("Segment Analysis")

    ws['A1'] = "SEGMENT PERFORMANCE BY MONTH"
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['primary'])

    segment_monthly = aggregate_by_segment_and_month(campaign_df)

    # Get unique months in order
    month_order = campaign_df.groupby('Month_Label')['Date'].min().sort_values()
    months = list(month_order.index)

    segments = ['Branded', 'Competitor', 'Non-Branded']

    # Spend by Segment
    ws['A3'] = "SPEND BY SEGMENT"
    ws['A3'].font = Font(bold=True, size=12, color=COLORS['primary'])

    row = 4
    headers = ["Segment"] + months
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        apply_header_style(cell)

    for seg in segments:
        row += 1
        ws.cell(row=row, column=1, value=seg).font = Font(bold=True)
        for col, month in enumerate(months, 2):
            val = segment_monthly[(segment_monthly['Segment'] == seg) & (segment_monthly['Month_Label'] == month)]['Spend'].sum()
            ws.cell(row=row, column=col, value=val).number_format = '$#,##0'

    # Sales by Segment
    row += 3
    ws.cell(row=row, column=1, value="SALES BY SEGMENT").font = Font(bold=True, size=12, color=COLORS['jn'])

    row += 1
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        apply_header_style(cell, 'jn')

    for seg in segments:
        row += 1
        ws.cell(row=row, column=1, value=seg).font = Font(bold=True)
        for col, month in enumerate(months, 2):
            val = segment_monthly[(segment_monthly['Segment'] == seg) & (segment_monthly['Month_Label'] == month)]['Sales'].sum()
            ws.cell(row=row, column=col, value=val).number_format = '$#,##0'

    # ROAS by Segment
    row += 3
    ws.cell(row=row, column=1, value="ROAS BY SEGMENT").font = Font(bold=True, size=12, color=COLORS['branded'])

    row += 1
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        apply_header_style(cell, 'branded')

    for seg in segments:
        row += 1
        ws.cell(row=row, column=1, value=seg).font = Font(bold=True)
        for col, month in enumerate(months, 2):
            data = segment_monthly[(segment_monthly['Segment'] == seg) & (segment_monthly['Month_Label'] == month)]
            val = data['ROAS'].values[0] if len(data) > 0 else 0
            ws.cell(row=row, column=col, value=val).number_format = '0.00'

    ws.column_dimensions['A'].width = 14
    for i in range(2, len(months) + 2):
        ws.column_dimensions[get_column_letter(i)].width = 12

def create_portfolio_sheet(wb, campaign_df):
    """Create Portfolio Analysis sheet."""
    ws = wb.create_sheet("Portfolio Analysis")

    ws['A1'] = "PORTFOLIO PERFORMANCE BY MONTH"
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['primary'])

    portfolio_monthly = aggregate_by_portfolio_and_month(campaign_df)

    month_order = campaign_df.groupby('Month_Label')['Date'].min().sort_values()
    months = list(month_order.index)

    portfolios = ['JN', 'Non-JN']

    # Spend by Portfolio
    ws['A3'] = "SPEND BY PORTFOLIO"
    ws['A3'].font = Font(bold=True, size=12, color=COLORS['primary'])

    row = 4
    headers = ["Portfolio"] + months + ["Total"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        apply_header_style(cell)

    for port in portfolios:
        row += 1
        ws.cell(row=row, column=1, value=port).font = Font(bold=True)
        total = 0
        for col, month in enumerate(months, 2):
            val = portfolio_monthly[(portfolio_monthly['Portfolio_Type'] == port) & (portfolio_monthly['Month_Label'] == month)]['Spend'].sum()
            ws.cell(row=row, column=col, value=val).number_format = '$#,##0'
            total += val
        ws.cell(row=row, column=len(months)+2, value=total).number_format = '$#,##0'

    # Sales by Portfolio
    row += 3
    ws.cell(row=row, column=1, value="SALES BY PORTFOLIO").font = Font(bold=True, size=12, color=COLORS['jn'])

    row += 1
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        apply_header_style(cell, 'jn')

    for port in portfolios:
        row += 1
        ws.cell(row=row, column=1, value=port).font = Font(bold=True)
        total = 0
        for col, month in enumerate(months, 2):
            val = portfolio_monthly[(portfolio_monthly['Portfolio_Type'] == port) & (portfolio_monthly['Month_Label'] == month)]['Sales'].sum()
            ws.cell(row=row, column=col, value=val).number_format = '$#,##0'
            total += val
        ws.cell(row=row, column=len(months)+2, value=total).number_format = '$#,##0'

    # ROAS by Portfolio
    row += 3
    ws.cell(row=row, column=1, value="ROAS BY PORTFOLIO").font = Font(bold=True, size=12, color=COLORS['branded'])

    row += 1
    for col, h in enumerate(headers[:len(headers)-1], 1):  # No total for ROAS
        cell = ws.cell(row=row, column=col, value=h if h != "Total" else "")
        apply_header_style(cell, 'branded')

    for port in portfolios:
        row += 1
        ws.cell(row=row, column=1, value=port).font = Font(bold=True)
        for col, month in enumerate(months, 2):
            data = portfolio_monthly[(portfolio_monthly['Portfolio_Type'] == port) & (portfolio_monthly['Month_Label'] == month)]
            val = data['ROAS'].values[0] if len(data) > 0 else 0
            ws.cell(row=row, column=col, value=val).number_format = '0.00'

    ws.column_dimensions['A'].width = 12
    for i in range(2, len(months) + 3):
        ws.column_dimensions[get_column_letter(i)].width = 12

def create_organic_sheet(wb, campaign_df, business_df, monthly_data):
    """Create Organic vs Paid Analysis sheet."""
    ws = wb.create_sheet("Organic vs Paid")

    ws['A1'] = "ORGANIC VS PAID ANALYSIS"
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['competitor'])

    if business_df is None or 'Total_Sales' not in monthly_data.columns:
        ws['A3'] = "Business Report data not available"
        return

    # Summary
    total_sales = business_df['Total_Sales'].sum()
    ad_sales = campaign_df['Sales'].sum()
    organic_sales = max(0, total_sales - ad_sales)

    ws['A3'] = "OVERALL SUMMARY"
    ws['A3'].font = Font(bold=True, size=12, color=COLORS['primary'])

    summary = [
        ("Total Sales", total_sales),
        ("Ad Sales", ad_sales),
        ("Organic Sales", organic_sales),
        ("Ad %", ad_sales/total_sales if total_sales > 0 else 0),
        ("Organic %", organic_sales/total_sales if total_sales > 0 else 0),
    ]

    row = 4
    for col, (label, value) in enumerate(summary, 1):
        cell = ws.cell(row=row, column=col, value=label)
        apply_header_style(cell, 'competitor')

    row = 5
    for col, (label, value) in enumerate(summary, 1):
        cell = ws.cell(row=row, column=col, value=value)
        if '%' in label:
            cell.number_format = '0.0%'
        else:
            cell.number_format = '$#,##0'
        cell.font = Font(bold=True, size=14)

    # Monthly breakdown
    row = 8
    ws.cell(row=row, column=1, value="MONTHLY BREAKDOWN").font = Font(bold=True, size=12, color=COLORS['primary'])

    row = 9
    headers = ["Month", "Total Sales", "Ad Sales", "Organic Sales", "Ad %", "Organic %", "TACOS"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        apply_header_style(cell, 'competitor')

    for idx, mrow in monthly_data.iterrows():
        row += 1
        ws.cell(row=row, column=1, value=mrow['Month_Label'])
        ws.cell(row=row, column=2, value=mrow.get('Total_Sales', 0)).number_format = '$#,##0'
        ws.cell(row=row, column=3, value=mrow['Sales']).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=mrow.get('Organic_Sales', 0)).number_format = '$#,##0'

        total = mrow.get('Total_Sales', 0)
        ad_pct = mrow['Sales'] / total if total > 0 else 0
        org_pct = mrow.get('Organic_Sales', 0) / total if total > 0 else 0

        ws.cell(row=row, column=5, value=ad_pct).number_format = '0.0%'
        ws.cell(row=row, column=6, value=org_pct).number_format = '0.0%'
        ws.cell(row=row, column=7, value=mrow.get('TACOS', 0)/100).number_format = '0.0%'

    for i in range(1, 8):
        ws.column_dimensions[get_column_letter(i)].width = 14

def create_raw_data_sheet(wb, df, sheet_name, key_columns):
    """Create a sheet with raw data for reference."""
    ws = wb.create_sheet(sheet_name)

    # Select key columns
    df_export = df[key_columns].copy()

    # Write headers
    for col, header in enumerate(df_export.columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell)

    # Write data (limit to first 10000 rows for Excel Online compatibility)
    max_rows = min(len(df_export), 10000)
    for row_idx in range(max_rows):
        for col_idx, col_name in enumerate(df_export.columns, 1):
            value = df_export.iloc[row_idx][col_name]
            cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)

            # Format dates
            if isinstance(value, pd.Timestamp):
                cell.number_format = 'YYYY-MM-DD'

    if len(df_export) > max_rows:
        ws.cell(row=max_rows + 3, column=1, value=f"Note: Showing first {max_rows:,} of {len(df_export):,} rows")

    for i in range(1, len(df_export.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

# ============================================================================
# MAIN FUNCTION
# ============================================================================

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    campaign_path = os.path.join(script_dir, CAMPAIGN_FILE)
    business_path = os.path.join(script_dir, BUSINESS_FILE)
    output_path = os.path.join(script_dir, OUTPUT_FILE)

    # Load data
    campaign_df = load_campaign_data(campaign_path)

    business_df = None
    if os.path.exists(business_path):
        business_df = load_business_data(business_path)
    else:
        print(f"Warning: Business report not found at {business_path}")

    # Aggregate data
    print("\nAggregating data...")
    monthly_data = aggregate_by_month(campaign_df, business_df)
    weekly_data = aggregate_by_week(campaign_df)

    print(f"  Monthly periods: {len(monthly_data)}")
    print(f"  Weekly periods: {len(weekly_data)}")

    # Create workbook
    print("\nGenerating Excel report...")
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create sheets
    create_summary_sheet(wb, campaign_df, business_df, monthly_data)
    create_monthly_sheet(wb, monthly_data)
    create_weekly_sheet(wb, weekly_data)
    create_segment_sheet(wb, campaign_df)
    create_portfolio_sheet(wb, campaign_df)

    if business_df is not None:
        create_organic_sheet(wb, campaign_df, business_df, monthly_data)

    # Raw data sheets (limited for Excel Online)
    campaign_cols = ['Date', 'Portfolio name', 'Campaign Name', 'Spend', 'Sales', 'Orders',
                     'Clicks', 'Impressions', 'Portfolio_Type', 'Segment', 'Month_Label']
    create_raw_data_sheet(wb, campaign_df, "Campaign Data", campaign_cols)

    if business_df is not None:
        business_cols = ['Date', 'Total_Sales', 'Units', 'Sessions', 'Month_Label']
        create_raw_data_sheet(wb, business_df, "Business Data", business_cols)

    # Save
    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")
    print("\nSheets created:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

if __name__ == "__main__":
    main()
