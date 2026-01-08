#!/usr/bin/env python3
"""
Campaign Performance Report Generator - Excel Version
Replicates the HTML Campaign Report Generator functionality in Excel
with data validation dropdowns for toggles and full interactivity.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, NamedStyle,
    numbers, Color
)
from openpyxl.formatting.rule import (
    FormulaRule, ColorScaleRule, DataBarRule, CellIsRule
)
from openpyxl.chart import (
    LineChart, PieChart, BarChart, Reference
)
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from datetime import datetime, timedelta
import string


# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

# Colors matching HTML version
COLORS = {
    'primary': '1E3A5F',
    'secondary': '2E86AB',
    'jn': '2F5496',
    'non_jn': '0EA5E9',
    'branded': '22C55E',
    'competitor': 'F97316',
    'non_branded': '8B5CF6',
    'text': '1F2937',
    'muted': '6B7280',
    'light_gray': 'F3F4F6',
    'border': 'D1D5DB',
    'white': 'FFFFFF',
    'positive': '22C55E',
    'negative': 'EF4444',
}

# Border styles
thin_border = Border(
    left=Side(style='thin', color=COLORS['border']),
    right=Side(style='thin', color=COLORS['border']),
    top=Side(style='thin', color=COLORS['border']),
    bottom=Side(style='thin', color=COLORS['border'])
)

medium_border = Border(
    left=Side(style='medium', color=COLORS['primary']),
    right=Side(style='medium', color=COLORS['primary']),
    top=Side(style='medium', color=COLORS['primary']),
    bottom=Side(style='medium', color=COLORS['primary'])
)


def create_header_style(color='primary'):
    """Create a header style with specified color."""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color=COLORS[color], end_color=COLORS[color], fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thin_border
    }


def apply_header_style(cell, color='primary'):
    """Apply header styling to a cell."""
    style = create_header_style(color)
    cell.font = style['font']
    cell.fill = style['fill']
    cell.alignment = style['alignment']
    cell.border = style['border']


def apply_metric_card_style(cell, is_value=False):
    """Apply metric card styling."""
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border
    if is_value:
        cell.font = Font(bold=True, size=16, color=COLORS['text'])
    else:
        cell.font = Font(size=9, color=COLORS['muted'])


def apply_currency_format(cell):
    """Apply currency format to cell."""
    cell.number_format = '$#,##0'


def apply_percent_format(cell, decimals=1):
    """Apply percentage format to cell."""
    cell.number_format = f'0.{"0" * decimals}%'


def apply_decimal_format(cell, decimals=2):
    """Apply decimal format to cell."""
    cell.number_format = f'0.{"0" * decimals}'


def apply_change_format(ws, cell_ref, value_ref=None):
    """Apply conditional formatting for positive/negative changes."""
    # Green for positive
    ws.conditional_formatting.add(
        cell_ref,
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
                   font=Font(color=COLORS['positive']))
    )
    # Red for negative
    ws.conditional_formatting.add(
        cell_ref,
        CellIsRule(operator='lessThan', formula=['0'],
                   fill=PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid'),
                   font=Font(color=COLORS['negative']))
    )


# ============================================================================
# WORKSHEET CREATION FUNCTIONS
# ============================================================================

def create_instructions_sheet(wb):
    """Create the Instructions sheet."""
    ws = wb.create_sheet("Instructions", 0)

    # Title
    ws['A1'] = "Campaign Performance Report - Excel Version"
    ws['A1'].font = Font(bold=True, size=20, color=COLORS['primary'])
    ws.merge_cells('A1:G1')

    ws['A2'] = "Replicates HTML Campaign Report Generator with Excel-native controls"
    ws['A2'].font = Font(size=12, color=COLORS['muted'])
    ws.merge_cells('A2:G2')

    # Section 1: Getting Started
    row = 4
    ws[f'A{row}'] = "1. GETTING STARTED"
    ws[f'A{row}'].font = Font(bold=True, size=14, color=COLORS['primary'])
    row += 1

    instructions = [
        "1.1 Import your Sponsored Products Campaign Report CSV into the 'Campaign Data' sheet",
        "1.2 Import your Business Report CSV into the 'Business Data' sheet (required for TACOS)",
        "1.3 Set your date range in the 'Settings' sheet",
        "1.4 Use the Portfolio dropdown to switch between Overall, JN, and Non-JN views",
        "1.5 Use the Time Period dropdown to switch between Daily, Weekly, and Monthly views",
    ]

    for instr in instructions:
        ws[f'A{row}'] = instr
        ws[f'A{row}'].font = Font(size=11)
        row += 1

    row += 1

    # Section 2: Sheet Overview
    ws[f'A{row}'] = "2. SHEET OVERVIEW"
    ws[f'A{row}'].font = Font(bold=True, size=14, color=COLORS['primary'])
    row += 1

    sheets = [
        ("Dashboard", "Main control panel with KPIs and portfolio selection"),
        ("Executive Summary", "Key metrics with month-over-month comparisons"),
        ("Segment Performance", "Branded/Competitor/Non-Branded breakdown"),
        ("Performance Trends", "Daily/Weekly/Monthly trend analysis"),
        ("Monthly Analysis", "Month-over-month performance data"),
        ("Weekly Analysis", "Week-over-week performance data"),
        ("Organic vs Paid", "Total Sales breakdown (requires Business Report)"),
        ("JN Portfolio", "JN-specific metrics and segments"),
        ("Non-JN Portfolio", "Non-JN-specific metrics and segments"),
        ("Pivot - Portfolio", "Cross-tabulation by portfolio and time"),
        ("Pivot - Segment", "Cross-tabulation by segment and time"),
        ("Campaign Data", "Raw campaign data input (paste your CSV here)"),
        ("Business Data", "Raw business report input (paste your CSV here)"),
        ("Settings", "Configuration and dropdown values"),
    ]

    ws[f'A{row}'] = "Sheet Name"
    ws[f'B{row}'] = "Description"
    apply_header_style(ws[f'A{row}'])
    apply_header_style(ws[f'B{row}'])
    row += 1

    for sheet_name, desc in sheets:
        ws[f'A{row}'] = sheet_name
        ws[f'B{row}'] = desc
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    row += 1

    # Section 3: Controls
    ws[f'A{row}'] = "3. INTERACTIVE CONTROLS"
    ws[f'A{row}'].font = Font(bold=True, size=14, color=COLORS['primary'])
    row += 1

    controls = [
        ("Portfolio Toggle", "Overall | JN | Non-JN", "Filters all metrics by portfolio"),
        ("Time Period", "Daily | Weekly | Monthly", "Aggregates trend data by time period"),
        ("Start Date", "Date picker", "Beginning of analysis period"),
        ("End Date", "Date picker", "End of analysis period"),
    ]

    ws[f'A{row}'] = "Control"
    ws[f'B{row}'] = "Options"
    ws[f'C{row}'] = "Effect"
    apply_header_style(ws[f'A{row}'])
    apply_header_style(ws[f'B{row}'])
    apply_header_style(ws[f'C{row}'])
    row += 1

    for ctrl, opts, effect in controls:
        ws[f'A{row}'] = ctrl
        ws[f'B{row}'] = opts
        ws[f'C{row}'] = effect
        row += 1

    row += 1

    # Section 4: Metrics
    ws[f'A{row}'] = "4. KEY METRICS DEFINITIONS"
    ws[f'A{row}'].font = Font(bold=True, size=14, color=COLORS['primary'])
    row += 1

    metrics = [
        ("Ad Spend", "Total advertising spend"),
        ("Ad Sales", "Sales attributed to advertising (7-day attribution)"),
        ("ROAS", "Return on Ad Spend = Ad Sales / Ad Spend"),
        ("ACoS", "Advertising Cost of Sale = Ad Spend / Ad Sales × 100%"),
        ("TACOS", "Total Advertising Cost of Sale = Ad Spend / Total Sales × 100%"),
        ("CVR", "Conversion Rate = Orders / Clicks × 100%"),
        ("CPC", "Cost Per Click = Ad Spend / Clicks"),
        ("CTR", "Click-Through Rate = Clicks / Impressions × 100%"),
        ("Organic Sales", "Total Sales - Ad Sales"),
    ]

    ws[f'A{row}'] = "Metric"
    ws[f'B{row}'] = "Definition"
    apply_header_style(ws[f'A{row}'])
    apply_header_style(ws[f'B{row}'])
    row += 1

    for metric, defn in metrics:
        ws[f'A{row}'] = metric
        ws[f'B{row}'] = defn
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    row += 2
    ws[f'A{row}'] = "Made by Krell for Piping Rock"
    ws[f'A{row}'].font = Font(italic=True, color=COLORS['muted'])

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 40


def create_settings_sheet(wb):
    """Create the Settings sheet with dropdown values and configuration."""
    ws = wb.create_sheet("Settings")

    # Title
    ws['A1'] = "Report Settings & Configuration"
    ws['A1'].font = Font(bold=True, size=16, color=COLORS['primary'])
    ws.merge_cells('A1:D1')

    # Section 1: Global Controls
    ws['A3'] = "GLOBAL CONTROLS"
    ws['A3'].font = Font(bold=True, size=12, color=COLORS['primary'])

    ws['A4'] = "Portfolio:"
    ws['B4'] = "Overall"
    ws['B4'].font = Font(bold=True, size=12)
    ws['B4'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

    ws['A5'] = "Time Period:"
    ws['B5'] = "Weekly"
    ws['B5'].font = Font(bold=True, size=12)
    ws['B5'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

    ws['A6'] = "Start Date:"
    ws['B6'] = datetime.now() - timedelta(days=90)
    ws['B6'].number_format = 'YYYY-MM-DD'
    ws['B6'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

    ws['A7'] = "End Date:"
    ws['B7'] = datetime.now()
    ws['B7'].number_format = 'YYYY-MM-DD'
    ws['B7'].fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

    # Section 2: Dropdown Lists (for data validation sources)
    ws['A10'] = "DROPDOWN OPTIONS (DO NOT MODIFY)"
    ws['A10'].font = Font(bold=True, size=12, color=COLORS['primary'])

    ws['A11'] = "Portfolio Options:"
    ws['B11'] = "Overall"
    ws['C11'] = "JN"
    ws['D11'] = "Non-JN"

    ws['A12'] = "Time Period Options:"
    ws['B12'] = "Daily"
    ws['C12'] = "Weekly"
    ws['D12'] = "Monthly"

    ws['A13'] = "Segment Options:"
    ws['B13'] = "Branded"
    ws['C13'] = "Competitor"
    ws['D13'] = "Non-Branded"

    # Add data validation to control cells (using simple list for Excel Online compatibility)
    portfolio_dv = DataValidation(
        type="list",
        formula1='"Overall,JN,Non-JN"',
        allow_blank=False
    )
    portfolio_dv.error = "Please select from dropdown"
    portfolio_dv.prompt = "Select portfolio view"
    ws.add_data_validation(portfolio_dv)
    portfolio_dv.add(ws['B4'])

    period_dv = DataValidation(
        type="list",
        formula1='"Daily,Weekly,Monthly"',
        allow_blank=False
    )
    period_dv.error = "Please select from dropdown"
    period_dv.prompt = "Select time period"
    ws.add_data_validation(period_dv)
    period_dv.add(ws['B5'])

    # Section 3: Calculated Constants
    ws['A16'] = "CALCULATED VALUES"
    ws['A16'].font = Font(bold=True, size=12, color=COLORS['primary'])

    ws['A17'] = "Report Generated:"
    ws['B17'] = datetime.now()
    ws['B17'].number_format = 'YYYY-MM-DD HH:MM'

    ws['A18'] = "Date Range Days:"
    ws['B18'] = '=B7-B6'

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15


def create_campaign_data_sheet(wb):
    """Create the Campaign Data input sheet with sample headers."""
    ws = wb.create_sheet("Campaign Data")

    # Instructions at top
    ws['A1'] = "PASTE YOUR SPONSORED PRODUCTS CAMPAIGN REPORT DATA BELOW"
    ws['A1'].font = Font(bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells('A1:R1')

    ws['A2'] = "Format: Copy all data including headers from your CSV export"
    ws['A2'].font = Font(italic=True, color=COLORS['muted'])

    # Expected headers (row 4)
    headers = [
        "Date", "Portfolio name", "Campaign Name", "Ad Group Name",
        "Impressions", "Clicks", "Spend", "7 Day Total Sales",
        "7 Day Total Orders (#)", "7 Day Conversion Rate",
        "Cost Per Click (CPC)", "Click-Thru Rate (CTR)",
        "7 Day Advertised SKU Sales", "7 Day Other SKU Sales"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell)

    # Add calculated column headers (formulas will be added by user after pasting data)
    calc_headers = [
        ("O4", "Portfolio Type"),
        ("P4", "Segment"),
        ("Q4", "Week"),
        ("R4", "Month"),
    ]

    for cell_ref, header in calc_headers:
        cell = ws[cell_ref]
        cell.value = header
        apply_header_style(cell)

    # Add formula instructions row
    ws['O3'] = "Formula: =IF(ISNUMBER(SEARCH(\"JN\",B5)),\"JN\",\"Non-JN\")"
    ws['O3'].font = Font(size=8, italic=True, color=COLORS['muted'])
    ws['P3'] = "Formula: =IF(ISNUMBER(SEARCH(\"branded\",C5)),\"Branded\",IF(OR(ISNUMBER(SEARCH(\" pat \",C5)),ISNUMBER(SEARCH(\"- pat -\",C5))),\"Competitor\",\"Non-Branded\"))"
    ws['P3'].font = Font(size=8, italic=True, color=COLORS['muted'])
    ws['Q3'] = "Formula: =TEXT(A5,\"YYYY\")\"-W\"&TEXT(WEEKNUM(A5),\"00\")"
    ws['Q3'].font = Font(size=8, italic=True, color=COLORS['muted'])
    ws['R3'] = "Formula: =TEXT(A5,\"MMM YYYY\")"
    ws['R3'].font = Font(size=8, italic=True, color=COLORS['muted'])

    # Set column widths
    widths = [12, 20, 40, 30, 12, 10, 12, 14, 14, 12, 12, 12, 14, 14, 12, 14, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Create as Excel Table for easier filtering
    ws['A3'] = "Data Table:"
    ws['A3'].font = Font(bold=True, size=11, color=COLORS['secondary'])


def create_business_data_sheet(wb):
    """Create the Business Data input sheet with sample headers."""
    ws = wb.create_sheet("Business Data")

    # Instructions at top
    ws['A1'] = "PASTE YOUR BUSINESS REPORT DATA BELOW"
    ws['A1'].font = Font(bold=True, size=12, color=COLORS['competitor'])
    ws.merge_cells('A1:J1')

    ws['A2'] = "Required for TACOS calculation and Organic vs Paid analysis"
    ws['A2'].font = Font(italic=True, color=COLORS['muted'])

    # Expected headers (row 4)
    headers = [
        "Date", "Ordered Product Sales", "Units Ordered",
        "Sessions", "Page Views", "Session Percentage",
        "Unit Session Percentage"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        apply_header_style(cell, 'competitor')

    # Add calculated column headers (formulas will be added by user after pasting data)
    calc_headers = [
        ("H4", "Week"),
        ("I4", "Month"),
    ]

    for cell_ref, header in calc_headers:
        cell = ws[cell_ref]
        cell.value = header
        apply_header_style(cell, 'competitor')

    # Add formula instructions row
    ws['H3'] = "Formula: =TEXT(A5,\"YYYY\")\"-W\"&TEXT(WEEKNUM(A5),\"00\")"
    ws['H3'].font = Font(size=8, italic=True, color=COLORS['muted'])
    ws['I3'] = "Formula: =TEXT(A5,\"MMM YYYY\")"
    ws['I3'].font = Font(size=8, italic=True, color=COLORS['muted'])

    # Set column widths
    widths = [12, 18, 14, 12, 12, 16, 18, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_dashboard_sheet(wb):
    """Create the main Dashboard with KPIs and controls."""
    ws = wb.create_sheet("Dashboard", 1)

    # Title
    ws['A1'] = "CAMPAIGN PERFORMANCE REPORT"
    ws['A1'].font = Font(bold=True, size=24, color=COLORS['primary'])
    ws.merge_cells('A1:H1')

    ws['A2'] = "Executive Dashboard"
    ws['A2'].font = Font(size=14, color=COLORS['muted'])
    ws.merge_cells('A2:H2')

    # Date range display
    ws['A3'] = "Date Range:"
    ws['B3'] = "=Settings!B6"
    ws['B3'].number_format = 'MMM D, YYYY'
    ws['C3'] = "to"
    ws['D3'] = "=Settings!B7"
    ws['D3'].number_format = 'MMM D, YYYY'

    # Control Panel
    ws['A5'] = "CONTROLS"
    ws['A5'].font = Font(bold=True, size=12, color=COLORS['primary'])
    ws['A5'].fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')
    ws.merge_cells('A5:D5')

    ws['A6'] = "Portfolio:"
    ws['B6'] = "=Settings!B4"
    ws['B6'].font = Font(bold=True, size=14)
    ws['B6'].fill = PatternFill(start_color=COLORS['jn'], end_color=COLORS['jn'], fill_type='solid')
    ws['B6'].font = Font(bold=True, size=14, color='FFFFFF')

    ws['C6'] = "Time Period:"
    ws['D6'] = "=Settings!B5"
    ws['D6'].font = Font(bold=True, size=14)
    ws['D6'].fill = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')
    ws['D6'].font = Font(bold=True, size=14, color='FFFFFF')

    # Headline Metric (Ad Sales)
    ws['A8'] = "TOTAL AD SALES"
    ws['A8'].font = Font(size=12, color='FFFFFF')
    ws['A8'].fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    ws['A8'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A8:D8')

    # Formula to calculate Ad Sales based on portfolio selection
    ws['A9'] = '=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!O:O,IF(Settings!B4="Overall","*",IF(Settings!B4="JN","JN","Non-JN")))'
    ws['A9'].number_format = '$#,##0'
    ws['A9'].font = Font(bold=True, size=36, color='FFFFFF')
    ws['A9'].fill = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')
    ws['A9'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A9:D9')

    # 4 KPI Cards
    kpi_start_row = 11
    kpis = [
        ("AD SPEND", "G", "$#,##0"),
        ("AD SALES", "H", "$#,##0"),
        ("ROAS", None, "0.00"),  # Calculated
        ("TACOS", None, "0.0%"),  # Calculated
    ]

    for col, (label, col_letter, fmt) in enumerate(kpis, 0):
        start_col = col * 2 + 1

        # Label
        cell = ws.cell(row=kpi_start_row, column=start_col, value=label)
        cell.font = Font(size=9, color=COLORS['muted'])
        cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=kpi_start_row, start_column=start_col, end_row=kpi_start_row, end_column=start_col+1)

        # Value
        value_cell = ws.cell(row=kpi_start_row+1, column=start_col)
        if col_letter:
            value_cell.value = f'=SUMIFS(\'Campaign Data\'!{col_letter}:{col_letter},\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7)'
        elif label == "ROAS":
            value_cell.value = '=IF(A12>0,B12/A12,0)'
        elif label == "TACOS":
            value_cell.value = '=IF(SUMIFS(\'Business Data\'!B:B,\'Business Data\'!A:A,">="&Settings!B6,\'Business Data\'!A:A,"<="&Settings!B7)>0,A12/SUMIFS(\'Business Data\'!B:B,\'Business Data\'!A:A,">="&Settings!B6,\'Business Data\'!A:A,"<="&Settings!B7),0)'

        value_cell.number_format = fmt
        value_cell.font = Font(bold=True, size=24, color=COLORS['text'])
        value_cell.alignment = Alignment(horizontal='center')
        value_cell.border = thin_border
        ws.merge_cells(start_row=kpi_start_row+1, start_column=start_col, end_row=kpi_start_row+1, end_column=start_col+1)

    # Organic Delta Section
    ws['A14'] = "SALES BREAKDOWN"
    ws['A14'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A14'].fill = PatternFill(start_color=COLORS['branded'], end_color=COLORS['branded'], fill_type='solid')
    ws.merge_cells('A14:H14')

    breakdown_labels = ["Total Sales", "Ad Sales", "Ad %", "Organic Sales", "Organic %"]
    for col, label in enumerate(breakdown_labels, 1):
        cell = ws.cell(row=15, column=col, value=label)
        cell.font = Font(bold=True, size=10, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLORS['branded'], end_color=COLORS['branded'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Formulas for breakdown
    ws['A16'] = '=SUMIFS(\'Business Data\'!B:B,\'Business Data\'!A:A,">="&Settings!B6,\'Business Data\'!A:A,"<="&Settings!B7)'
    ws['A16'].number_format = '$#,##0'
    ws['B16'] = '=B12'  # Reference to Ad Sales KPI
    ws['B16'].number_format = '$#,##0'
    ws['C16'] = '=IF(A16>0,B16/A16,0)'
    ws['C16'].number_format = '0.0%'
    ws['D16'] = '=A16-B16'
    ws['D16'].number_format = '$#,##0'
    ws['E16'] = '=IF(A16>0,D16/A16,0)'
    ws['E16'].number_format = '0.0%'

    for col in range(1, 6):
        ws.cell(row=16, column=col).font = Font(bold=True, size=14)
        ws.cell(row=16, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=16, column=col).border = thin_border

    # Segment Performance Summary
    ws['A18'] = "SEGMENT PERFORMANCE (Current Period)"
    ws['A18'].font = Font(bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells('A18:E18')

    segment_headers = ["Segment", "Spend", "Sales", "ROAS", "ACoS"]
    for col, header in enumerate(segment_headers, 1):
        cell = ws.cell(row=19, column=col, value=header)
        apply_header_style(cell)

    segments = ["Branded", "Competitor", "Non-Branded"]
    for row, segment in enumerate(segments, 20):
        ws.cell(row=row, column=1, value=segment).font = Font(bold=True)

        # Spend formula
        spend_cell = ws.cell(row=row, column=2)
        spend_cell.value = f'=SUMIFS(\'Campaign Data\'!G:G,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!P:P,"{segment}")'
        spend_cell.number_format = '$#,##0'

        # Sales formula
        sales_cell = ws.cell(row=row, column=3)
        sales_cell.value = f'=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!P:P,"{segment}")'
        sales_cell.number_format = '$#,##0'

        # ROAS formula
        roas_cell = ws.cell(row=row, column=4)
        roas_cell.value = f'=IF(B{row}>0,C{row}/B{row},0)'
        roas_cell.number_format = '0.00'

        # ACoS formula
        acos_cell = ws.cell(row=row, column=5)
        acos_cell.value = f'=IF(C{row}>0,B{row}/C{row},0)'
        acos_cell.number_format = '0.0%'

    # Monthly Performance Summary
    ws['A25'] = "MONTHLY PERFORMANCE TREND"
    ws['A25'].font = Font(bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells('A25:G25')

    monthly_headers = ["Month", "Spend", "Sales", "ROAS", "MoM Spend %", "MoM Sales %", "MoM ROAS %"]
    for col, header in enumerate(monthly_headers, 1):
        cell = ws.cell(row=26, column=col, value=header)
        apply_header_style(cell)

    # Add placeholder rows for monthly data (will be populated by formulas)
    for row in range(27, 33):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col >= 5:
                apply_change_format(ws, cell.coordinate)

    # Set column widths
    widths = [15, 15, 15, 15, 15, 15, 15, 15]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Footer
    ws['A35'] = "Made by Krell for Piping Rock"
    ws['A35'].font = Font(italic=True, color=COLORS['muted'])
    ws.merge_cells('A35:D35')


def create_executive_summary_sheet(wb):
    """Create the Executive Summary sheet matching Page 1 of HTML report."""
    ws = wb.create_sheet("Executive Summary")

    # Title
    ws['A1'] = "EXECUTIVE SUMMARY"
    ws['A1'].font = Font(bold=True, size=18, color=COLORS['primary'])
    ws.merge_cells('A1:H1')

    # Portfolio and Date indicators
    ws['A2'] = "Portfolio:"
    ws['B2'] = "=Settings!B4"
    ws['B2'].font = Font(bold=True)
    ws['C2'] = "Period:"
    ws['D2'] = '=TEXT(Settings!B6,"MMM D, YYYY")&" - "&TEXT(Settings!B7,"MMM D, YYYY")'

    # Most Recent Complete Month indicator
    ws['A3'] = "Showing Most Recent Complete Month"
    ws['A3'].font = Font(italic=True, color=COLORS['muted'])

    # Headline Metric Box
    ws['A5'] = "AD SALES"
    ws['A5'].font = Font(bold=True, size=11, color='FFFFFF')
    ws['A5'].fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    ws['A5'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A5:D5')

    # Current month ad sales
    ws['A6'] = '=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!R:R,TEXT(EOMONTH(Settings!B7,-1)+1,"MMM YYYY"))'
    ws['A6'].number_format = '$#,##0'
    ws['A6'].font = Font(bold=True, size=32, color='FFFFFF')
    ws['A6'].fill = PatternFill(start_color=COLORS['secondary'], end_color=COLORS['secondary'], fill_type='solid')
    ws['A6'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A6:D6')

    # MoM change indicator
    ws['A7'] = '=IF(A6>0,IF(E6>0,(A6-E6)/E6,0),0)'  # Placeholder for MoM calculation
    ws['A7'].number_format = '0.0%'
    ws['A7'].font = Font(bold=True, size=12)
    ws['A7'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A7:D7')

    # Previous month value for comparison (hidden or in column E)
    ws['E6'] = '=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!R:R,TEXT(EOMONTH(Settings!B7,-2)+1,"MMM YYYY"))'
    ws['E6'].number_format = '$#,##0'
    ws.column_dimensions['E'].hidden = True

    # 4 KPI Cards Row
    kpi_row = 9
    kpis = [
        ("AD SPEND", "=SUMIFS('Campaign Data'!G:G,'Campaign Data'!R:R,TEXT(EOMONTH(Settings!B7,-1)+1,\"MMM YYYY\"))", "$#,##0"),
        ("AD SALES", "=SUMIFS('Campaign Data'!H:H,'Campaign Data'!R:R,TEXT(EOMONTH(Settings!B7,-1)+1,\"MMM YYYY\"))", "$#,##0"),
        ("ROAS", "=IF(A10>0,C10/A10,0)", "0.00\"x\""),
        ("TACOS", "=IF(SUMIFS('Business Data'!B:B,'Business Data'!I:I,TEXT(EOMONTH(Settings!B7,-1)+1,\"MMM YYYY\"))>0,A10/SUMIFS('Business Data'!B:B,'Business Data'!I:I,TEXT(EOMONTH(Settings!B7,-1)+1,\"MMM YYYY\")),0)", "0.0%"),
    ]

    for col, (label, formula, fmt) in enumerate(kpis):
        col_letter = get_column_letter(col * 2 + 1)
        next_letter = get_column_letter(col * 2 + 2)

        # Label
        label_cell = ws[f'{col_letter}{kpi_row}']
        label_cell.value = label
        label_cell.font = Font(size=9, color=COLORS['muted'])
        label_cell.alignment = Alignment(horizontal='center')

        # Value
        value_cell = ws[f'{col_letter}{kpi_row+1}']
        value_cell.value = formula
        value_cell.number_format = fmt
        value_cell.font = Font(bold=True, size=20)
        value_cell.alignment = Alignment(horizontal='center')
        value_cell.border = thin_border

    # Organic Delta Section
    ws['A13'] = "SALES BREAKDOWN (Monthly)"
    ws['A13'].font = Font(bold=True, size=11, color='FFFFFF')
    ws['A13'].fill = PatternFill(start_color=COLORS['branded'], end_color=COLORS['branded'], fill_type='solid')
    ws.merge_cells('A13:F13')

    breakdown_headers = ["Total Sales", "Ad Sales", "Organic Sales", "Ad %", "Organic %"]
    for col, header in enumerate(breakdown_headers, 1):
        cell = ws.cell(row=14, column=col, value=header)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=COLORS['light_gray'], end_color=COLORS['light_gray'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Values row
    ws['A15'] = '=SUMIFS(\'Business Data\'!B:B,\'Business Data\'!I:I,TEXT(EOMONTH(Settings!B7,-1)+1,"MMM YYYY"))'
    ws['A15'].number_format = '$#,##0'
    ws['B15'] = '=C10'  # Reference Ad Sales KPI
    ws['B15'].number_format = '$#,##0'
    ws['C15'] = '=MAX(0,A15-B15)'
    ws['C15'].number_format = '$#,##0'
    ws['D15'] = '=IF(A15>0,B15/A15,0)'
    ws['D15'].number_format = '0.0%'
    ws['E15'] = '=IF(A15>0,C15/A15,0)'
    ws['E15'].number_format = '0.0%'

    for col in range(1, 6):
        ws.cell(row=15, column=col).font = Font(bold=True, size=14)
        ws.cell(row=15, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=15, column=col).border = thin_border

    # Key Insights Section
    ws['A18'] = "KEY INSIGHTS"
    ws['A18'].font = Font(bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells('A18:H18')

    insights = [
        "1. Ad Sales performance vs previous month shows trend direction",
        "2. Best performing segment by ROAS indicates optimization opportunities",
        "3. TACOS trend indicates overall advertising efficiency vs total revenue",
    ]

    for i, insight in enumerate(insights, 19):
        ws[f'A{i}'] = insight
        ws[f'A{i}'].font = Font(size=11)
        ws.merge_cells(f'A{i}:H{i}')

    # Monthly Performance Table
    ws['A23'] = "MONTHLY PERFORMANCE"
    ws['A23'].font = Font(bold=True, size=12, color=COLORS['primary'])
    ws.merge_cells('A23:J23')

    monthly_headers = ["Metric", "Month -5", "Month -4", "Month -3", "Month -2", "Month -1", "Current"]
    for col, header in enumerate(monthly_headers, 1):
        cell = ws.cell(row=24, column=col, value=header)
        apply_header_style(cell)

    metrics = [
        ("Ad Spend", "$#,##0"),
        ("Ad Sales", "$#,##0"),
        ("ROAS", "0.00"),
        ("ACoS %", "0.0%"),
        ("Orders", "#,##0"),
        ("Clicks", "#,##0"),
        ("CVR %", "0.00%"),
    ]

    for row, (metric, fmt) in enumerate(metrics, 25):
        ws.cell(row=row, column=1, value=metric).font = Font(bold=True)
        for col in range(2, 8):
            cell = ws.cell(row=row, column=col)
            cell.number_format = fmt
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # Set column widths
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 14


def create_segment_performance_sheet(wb):
    """Create the Segment Performance sheet matching Page 2 of HTML report."""
    ws = wb.create_sheet("Segment Performance")

    # Title
    ws['A1'] = "SEGMENT PERFORMANCE"
    ws['A1'].font = Font(bold=True, size=18, color=COLORS['primary'])
    ws.merge_cells('A1:H1')

    ws['A2'] = "Branded / Competitor / Non-Branded Breakdown"
    ws['A2'].font = Font(size=12, color=COLORS['muted'])

    # Segment Performance Table
    ws['A4'] = "PERFORMANCE BY SEGMENT"
    ws['A4'].font = Font(bold=True, size=12, color=COLORS['primary'])

    headers = ["Segment", "Spend", "Spend %", "Sales", "Sales %", "ROAS", "ACoS", "MoM Spend", "MoM Sales"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_header_style(cell)

    segments = [
        ("Branded", COLORS['branded']),
        ("Competitor", COLORS['competitor']),
        ("Non-Branded", COLORS['non_branded']),
    ]

    for row, (segment, color) in enumerate(segments, 6):
        # Segment name with color indicator
        name_cell = ws.cell(row=row, column=1, value=segment)
        name_cell.font = Font(bold=True)
        name_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        name_cell.font = Font(bold=True, color='FFFFFF')

        # Spend
        ws.cell(row=row, column=2, value=f'=SUMIFS(\'Campaign Data\'!G:G,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!P:P,"{segment}")')
        ws.cell(row=row, column=2).number_format = '$#,##0'

        # Spend %
        ws.cell(row=row, column=3, value=f'=IF(SUM(B6:B8)>0,B{row}/SUM(B6:B8),0)')
        ws.cell(row=row, column=3).number_format = '0.0%'

        # Sales
        ws.cell(row=row, column=4, value=f'=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!P:P,"{segment}")')
        ws.cell(row=row, column=4).number_format = '$#,##0'

        # Sales %
        ws.cell(row=row, column=5, value=f'=IF(SUM(D6:D8)>0,D{row}/SUM(D6:D8),0)')
        ws.cell(row=row, column=5).number_format = '0.0%'

        # ROAS
        ws.cell(row=row, column=6, value=f'=IF(B{row}>0,D{row}/B{row},0)')
        ws.cell(row=row, column=6).number_format = '0.00'

        # ACoS
        ws.cell(row=row, column=7, value=f'=IF(D{row}>0,B{row}/D{row},0)')
        ws.cell(row=row, column=7).number_format = '0.0%'

        # MoM Change columns (placeholders)
        ws.cell(row=row, column=8).number_format = '0.0%'
        ws.cell(row=row, column=9).number_format = '0.0%'

        for col in range(1, 10):
            ws.cell(row=row, column=col).border = thin_border

    # Total row
    ws.cell(row=9, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=9, column=2, value="=SUM(B6:B8)").number_format = '$#,##0'
    ws.cell(row=9, column=3, value="100%")
    ws.cell(row=9, column=4, value="=SUM(D6:D8)").number_format = '$#,##0'
    ws.cell(row=9, column=5, value="100%")
    ws.cell(row=9, column=6, value="=IF(B9>0,D9/B9,0)").number_format = '0.00'
    ws.cell(row=9, column=7, value="=IF(D9>0,B9/D9,0)").number_format = '0.0%'

    for col in range(1, 10):
        ws.cell(row=9, column=col).border = thin_border
        ws.cell(row=9, column=col).font = Font(bold=True)

    # Segment by Portfolio breakdown
    ws['A12'] = "SEGMENT BY PORTFOLIO"
    ws['A12'].font = Font(bold=True, size=12, color=COLORS['primary'])

    portfolio_headers = ["Segment", "JN Spend", "JN Sales", "JN ROAS", "Non-JN Spend", "Non-JN Sales", "Non-JN ROAS"]
    for col, header in enumerate(portfolio_headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        apply_header_style(cell, 'jn' if 'JN' in header and 'Non' not in header else 'non_jn' if 'Non-JN' in header else 'primary')

    for row, (segment, color) in enumerate(segments, 14):
        ws.cell(row=row, column=1, value=segment).font = Font(bold=True)

        # JN metrics
        ws.cell(row=row, column=2, value=f'=SUMIFS(\'Campaign Data\'!G:G,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!P:P,"{segment}",\'Campaign Data\'!O:O,"JN")')
        ws.cell(row=row, column=2).number_format = '$#,##0'

        ws.cell(row=row, column=3, value=f'=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!P:P,"{segment}",\'Campaign Data\'!O:O,"JN")')
        ws.cell(row=row, column=3).number_format = '$#,##0'

        ws.cell(row=row, column=4, value=f'=IF(B{row}>0,C{row}/B{row},0)')
        ws.cell(row=row, column=4).number_format = '0.00'

        # Non-JN metrics
        ws.cell(row=row, column=5, value=f'=SUMIFS(\'Campaign Data\'!G:G,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!P:P,"{segment}",\'Campaign Data\'!O:O,"Non-JN")')
        ws.cell(row=row, column=5).number_format = '$#,##0'

        ws.cell(row=row, column=6, value=f'=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!P:P,"{segment}",\'Campaign Data\'!O:O,"Non-JN")')
        ws.cell(row=row, column=6).number_format = '$#,##0'

        ws.cell(row=row, column=7, value=f'=IF(E{row}>0,F{row}/E{row},0)')
        ws.cell(row=row, column=7).number_format = '0.00'

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border

    # Set column widths
    widths = [14, 14, 12, 14, 12, 10, 10, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_performance_trends_sheet(wb):
    """Create the Performance Trends sheet matching Page 3 of HTML report."""
    ws = wb.create_sheet("Performance Trends")

    # Title
    ws['A1'] = "PERFORMANCE TRENDS"
    ws['A1'].font = Font(bold=True, size=18, color=COLORS['primary'])
    ws.merge_cells('A1:L1')

    ws['A2'] = "Time Period:"
    ws['B2'] = "=Settings!B5"
    ws['B2'].font = Font(bold=True)
    ws['C2'] = "Portfolio:"
    ws['D2'] = "=Settings!B4"
    ws['D2'].font = Font(bold=True)

    # Weekly Trend Table
    ws['A4'] = "WEEKLY PERFORMANCE"
    ws['A4'].font = Font(bold=True, size=12, color=COLORS['primary'])

    weekly_headers = ["Week", "Spend", "Sales", "ROAS", "ACoS", "WoW Spend %", "WoW Sales %", "WoW ROAS %"]
    for col, header in enumerate(weekly_headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        apply_header_style(cell)

    # Add placeholder rows for weekly data
    for row in range(6, 18):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            if col in [2, 3]:
                cell.number_format = '$#,##0'
            elif col == 4:
                cell.number_format = '0.00'
            elif col >= 5:
                cell.number_format = '0.0%'

    # Apply conditional formatting to change columns
    for col in ['F', 'G', 'H']:
        apply_change_format(ws, f'{col}6:{col}17')

    # Monthly Trend Table
    ws['A20'] = "MONTHLY PERFORMANCE"
    ws['A20'].font = Font(bold=True, size=12, color=COLORS['primary'])

    monthly_headers = ["Month", "Spend", "Sales", "ROAS", "ACoS", "Orders", "Clicks", "CVR", "MoM Spend", "MoM Sales"]
    for col, header in enumerate(monthly_headers, 1):
        cell = ws.cell(row=21, column=col, value=header)
        apply_header_style(cell)

    # Add placeholder rows for monthly data
    for row in range(22, 34):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # Apply conditional formatting to change columns
    for col in ['I', 'J']:
        apply_change_format(ws, f'{col}22:{col}33')

    # Segment Trend Section
    ws['A36'] = "SEGMENT TRENDS BY MONTH"
    ws['A36'].font = Font(bold=True, size=12, color=COLORS['primary'])

    # Branded trends
    ws['A38'] = "Branded"
    ws['A38'].font = Font(bold=True, color=COLORS['branded'])

    # Competitor trends
    ws['A48'] = "Competitor"
    ws['A48'].font = Font(bold=True, color=COLORS['competitor'])

    # Non-Branded trends
    ws['A58'] = "Non-Branded"
    ws['A58'].font = Font(bold=True, color=COLORS['non_branded'])

    # Set column widths
    widths = [14, 14, 14, 10, 10, 12, 12, 10, 12, 12, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_monthly_analysis_sheet(wb):
    """Create detailed Month-over-Month analysis sheet."""
    ws = wb.create_sheet("Monthly Analysis")

    ws['A1'] = "MONTH-OVER-MONTH ANALYSIS"
    ws['A1'].font = Font(bold=True, size=18, color=COLORS['primary'])
    ws.merge_cells('A1:J1')

    # Headers
    headers = ["Month", "Spend", "Spend Chg", "Spend %", "Sales", "Sales Chg", "Sales %", "ROAS", "ROAS Chg", "ACoS"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)

    # Placeholder rows
    for row in range(4, 16):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            if col in [2, 3, 5, 6]:
                cell.number_format = '$#,##0'
            elif col in [4, 7, 10]:
                cell.number_format = '0.0%'
            elif col in [8, 9]:
                cell.number_format = '0.00'

    # Apply conditional formatting
    for col in ['C', 'D', 'F', 'G', 'I']:
        apply_change_format(ws, f'{col}4:{col}15')

    # Set column widths
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 14


def create_weekly_analysis_sheet(wb):
    """Create detailed Week-over-Week analysis sheet."""
    ws = wb.create_sheet("Weekly Analysis")

    ws['A1'] = "WEEK-OVER-WEEK ANALYSIS"
    ws['A1'].font = Font(bold=True, size=18, color=COLORS['primary'])
    ws.merge_cells('A1:K1')

    # Headers
    headers = ["Week", "Date Range", "Spend", "Spend Chg", "Spend %", "Sales", "Sales Chg", "Sales %", "ROAS", "ACoS", "Orders"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell)

    # Placeholder rows
    for row in range(4, 20):
        for col in range(1, 12):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

    # Apply conditional formatting
    for col in ['D', 'E', 'G', 'H']:
        apply_change_format(ws, f'{col}4:{col}19')

    # Set column widths
    widths = [12, 22, 14, 14, 10, 14, 14, 10, 10, 10, 10, 10]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def create_organic_vs_paid_sheet(wb):
    """Create Organic vs Paid analysis sheet."""
    ws = wb.create_sheet("Organic vs Paid")

    ws['A1'] = "ORGANIC VS PAID ANALYSIS"
    ws['A1'].font = Font(bold=True, size=18, color=COLORS['competitor'])
    ws.merge_cells('A1:I1')

    ws['A2'] = "Requires Business Report data for Total Sales"
    ws['A2'].font = Font(italic=True, color=COLORS['muted'])

    # Summary Section
    ws['A4'] = "PERIOD SUMMARY"
    ws['A4'].font = Font(bold=True, size=12, color=COLORS['primary'])

    summary_labels = ["Total Sales", "Ad Sales", "Organic Sales", "Ad %", "Organic %", "TACOS"]
    for col, label in enumerate(summary_labels, 1):
        cell = ws.cell(row=5, column=col, value=label)
        apply_header_style(cell, 'branded')

    # Formulas
    ws['A6'] = '=SUMIFS(\'Business Data\'!B:B,\'Business Data\'!A:A,">="&Settings!B6,\'Business Data\'!A:A,"<="&Settings!B7)'
    ws['A6'].number_format = '$#,##0'
    ws['B6'] = '=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7)'
    ws['B6'].number_format = '$#,##0'
    ws['C6'] = '=MAX(0,A6-B6)'
    ws['C6'].number_format = '$#,##0'
    ws['D6'] = '=IF(A6>0,B6/A6,0)'
    ws['D6'].number_format = '0.0%'
    ws['E6'] = '=IF(A6>0,C6/A6,0)'
    ws['E6'].number_format = '0.0%'
    ws['F6'] = '=IF(A6>0,SUMIFS(\'Campaign Data\'!G:G,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7)/A6,0)'
    ws['F6'].number_format = '0.0%'

    for col in range(1, 7):
        ws.cell(row=6, column=col).font = Font(bold=True, size=14)
        ws.cell(row=6, column=col).border = thin_border
        ws.cell(row=6, column=col).alignment = Alignment(horizontal='center')

    # Monthly breakdown
    ws['A9'] = "MONTHLY ORGANIC VS PAID"
    ws['A9'].font = Font(bold=True, size=12, color=COLORS['primary'])

    headers = ["Month", "Total Sales", "Ad Sales", "Organic Sales", "Organic %", "Organic Chg", "Ad Sales Chg", "TACOS"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        apply_header_style(cell, 'competitor')

    # Placeholder rows
    for row in range(11, 23):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            if col in [2, 3, 4]:
                cell.number_format = '$#,##0'
            elif col >= 5:
                cell.number_format = '0.0%'

    # Apply conditional formatting
    for col in ['F', 'G']:
        apply_change_format(ws, f'{col}11:{col}22')

    # Set column widths
    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 15


def create_portfolio_sheets(wb):
    """Create JN and Non-JN portfolio detail sheets."""
    portfolios = [
        ("JN Portfolio", "JN", 'jn'),
        ("Non-JN Portfolio", "Non-JN", 'non_jn'),
    ]

    for sheet_name, portfolio_filter, color_key in portfolios:
        ws = wb.create_sheet(sheet_name)

        ws['A1'] = f"{portfolio_filter} PORTFOLIO ANALYSIS"
        ws['A1'].font = Font(bold=True, size=18, color=COLORS[color_key])
        ws.merge_cells('A1:G1')

        # Summary metrics
        ws['A3'] = "PORTFOLIO SUMMARY"
        ws['A3'].font = Font(bold=True, size=12, color=COLORS['primary'])

        metrics = ["Spend", "Sales", "ROAS", "ACoS", "Orders", "Clicks", "CVR"]
        for col, metric in enumerate(metrics, 1):
            cell = ws.cell(row=4, column=col, value=metric)
            apply_header_style(cell, color_key)

        # Formulas
        ws['A5'] = f'=SUMIFS(\'Campaign Data\'!G:G,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!O:O,"{portfolio_filter}")'
        ws['A5'].number_format = '$#,##0'
        ws['B5'] = f'=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!O:O,"{portfolio_filter}")'
        ws['B5'].number_format = '$#,##0'
        ws['C5'] = '=IF(A5>0,B5/A5,0)'
        ws['C5'].number_format = '0.00'
        ws['D5'] = '=IF(B5>0,A5/B5,0)'
        ws['D5'].number_format = '0.0%'
        ws['E5'] = f'=SUMIFS(\'Campaign Data\'!I:I,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!O:O,"{portfolio_filter}")'
        ws['E5'].number_format = '#,##0'
        ws['F5'] = f'=SUMIFS(\'Campaign Data\'!F:F,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!O:O,"{portfolio_filter}")'
        ws['F5'].number_format = '#,##0'
        ws['G5'] = '=IF(F5>0,E5/F5,0)'
        ws['G5'].number_format = '0.00%'

        for col in range(1, 8):
            ws.cell(row=5, column=col).font = Font(bold=True, size=14)
            ws.cell(row=5, column=col).border = thin_border
            ws.cell(row=5, column=col).alignment = Alignment(horizontal='center')

        # Segment breakdown
        ws['A8'] = "SEGMENT BREAKDOWN"
        ws['A8'].font = Font(bold=True, size=12, color=COLORS['primary'])

        segment_headers = ["Segment", "Spend", "Sales", "ROAS", "ACoS"]
        for col, header in enumerate(segment_headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            apply_header_style(cell, color_key)

        segments = ["Branded", "Competitor", "Non-Branded"]
        for row, segment in enumerate(segments, 10):
            ws.cell(row=row, column=1, value=segment).font = Font(bold=True)

            ws.cell(row=row, column=2, value=f'=SUMIFS(\'Campaign Data\'!G:G,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!O:O,"{portfolio_filter}",\'Campaign Data\'!P:P,"{segment}")')
            ws.cell(row=row, column=2).number_format = '$#,##0'

            ws.cell(row=row, column=3, value=f'=SUMIFS(\'Campaign Data\'!H:H,\'Campaign Data\'!A:A,">="&Settings!B6,\'Campaign Data\'!A:A,"<="&Settings!B7,\'Campaign Data\'!O:O,"{portfolio_filter}",\'Campaign Data\'!P:P,"{segment}")')
            ws.cell(row=row, column=3).number_format = '$#,##0'

            ws.cell(row=row, column=4, value=f'=IF(B{row}>0,C{row}/B{row},0)')
            ws.cell(row=row, column=4).number_format = '0.00'

            ws.cell(row=row, column=5, value=f'=IF(C{row}>0,B{row}/C{row},0)')
            ws.cell(row=row, column=5).number_format = '0.0%'

            for col in range(1, 6):
                ws.cell(row=row, column=col).border = thin_border

        # Set column widths
        for i in range(1, 8):
            ws.column_dimensions[get_column_letter(i)].width = 14


def create_pivot_sheets(wb):
    """Create pivot-style analysis sheets."""
    # Pivot - Monthly by Portfolio
    ws = wb.create_sheet("Pivot - Portfolio")

    ws['A1'] = "SPEND BY MONTH & PORTFOLIO TYPE"
    ws['A1'].font = Font(bold=True, size=14, color=COLORS['primary'])

    headers = ["Month", "JN", "Non-JN", "Total"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        apply_header_style(cell, 'jn' if header == 'JN' else 'non_jn' if header == 'Non-JN' else 'primary')

    # Placeholder rows
    for row in range(4, 16):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col > 1:
                cell.number_format = '$#,##0'

    # AD SALES section
    ws['A18'] = "AD SALES BY MONTH & PORTFOLIO TYPE"
    ws['A18'].font = Font(bold=True, size=14, color=COLORS['jn'])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=20, column=col, value=header)
        apply_header_style(cell, 'jn' if header == 'JN' else 'non_jn' if header == 'Non-JN' else 'primary')

    for row in range(21, 33):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col > 1:
                cell.number_format = '$#,##0'

    # ROAS section
    ws['A35'] = "ROAS BY MONTH & PORTFOLIO TYPE"
    ws['A35'].font = Font(bold=True, size=14, color=COLORS['branded'])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=37, column=col, value=header)
        apply_header_style(cell, 'branded')

    for row in range(38, 50):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col > 1:
                cell.number_format = '0.00'

    for i in range(1, 5):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # Pivot - Monthly by Segment
    ws2 = wb.create_sheet("Pivot - Segment")

    ws2['A1'] = "SPEND BY MONTH & SEGMENT"
    ws2['A1'].font = Font(bold=True, size=14, color=COLORS['primary'])

    seg_headers = ["Month", "Branded", "Competitor", "Non-Branded", "Total"]
    for col, header in enumerate(seg_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        color = 'branded' if header == 'Branded' else 'competitor' if header == 'Competitor' else 'non_branded' if header == 'Non-Branded' else 'primary'
        apply_header_style(cell, color)

    for row in range(4, 16):
        for col in range(1, 6):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            if col > 1:
                cell.number_format = '$#,##0'

    # AD SALES by Segment
    ws2['A18'] = "AD SALES BY MONTH & SEGMENT"
    ws2['A18'].font = Font(bold=True, size=14, color=COLORS['jn'])

    for col, header in enumerate(seg_headers, 1):
        cell = ws2.cell(row=20, column=col, value=header)
        color = 'branded' if header == 'Branded' else 'competitor' if header == 'Competitor' else 'non_branded' if header == 'Non-Branded' else 'primary'
        apply_header_style(cell, color)

    for row in range(21, 33):
        for col in range(1, 6):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            if col > 1:
                cell.number_format = '$#,##0'

    # ROAS by Segment
    ws2['A35'] = "ROAS BY MONTH & SEGMENT"
    ws2['A35'].font = Font(bold=True, size=14, color=COLORS['branded'])

    for col, header in enumerate(seg_headers, 1):
        cell = ws2.cell(row=37, column=col, value=header)
        apply_header_style(cell, 'branded')

    for row in range(38, 50):
        for col in range(1, 6):
            cell = ws2.cell(row=row, column=col)
            cell.border = thin_border
            if col > 1:
                cell.number_format = '0.00'

    for i in range(1, 6):
        ws2.column_dimensions[get_column_letter(i)].width = 15


# ============================================================================
# MAIN FUNCTION
# ============================================================================

def create_campaign_report_workbook():
    """Create the complete Campaign Performance Report Excel workbook."""
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create all sheets in order
    create_instructions_sheet(wb)
    create_settings_sheet(wb)
    create_dashboard_sheet(wb)
    create_executive_summary_sheet(wb)
    create_segment_performance_sheet(wb)
    create_performance_trends_sheet(wb)
    create_monthly_analysis_sheet(wb)
    create_weekly_analysis_sheet(wb)
    create_organic_vs_paid_sheet(wb)
    create_portfolio_sheets(wb)
    create_pivot_sheets(wb)
    create_campaign_data_sheet(wb)
    create_business_data_sheet(wb)

    return wb


def main():
    """Main entry point."""
    print("Generating Campaign Performance Report Excel Template...")

    wb = create_campaign_report_workbook()

    # Save to file (v2 = Excel Online compatible)
    output_path = os.path.join(os.path.dirname(__file__), "Campaign_Report_Template_v2.xlsx")
    wb.save(output_path)

    print(f"Excel template saved to: {output_path}")
    print("\nThe workbook contains the following sheets:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

    print("\nInstructions:")
    print("1. Open the Excel file")
    print("2. Paste your Campaign Report CSV data into 'Campaign Data' sheet (starting row 5)")
    print("3. Paste your Business Report CSV data into 'Business Data' sheet (starting row 5)")
    print("4. Adjust date range in 'Settings' sheet")
    print("5. Use Portfolio and Time Period dropdowns to filter views")


if __name__ == "__main__":
    main()
